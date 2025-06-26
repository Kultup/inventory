from flask import Flask, render_template, request, redirect, url_for, flash, abort, send_from_directory, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from datetime import datetime, timedelta
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import os
import uuid
import random
import string
import re
import io
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import func, extract
import qrcode
import shutil
import zipfile
import json
from apscheduler.schedulers.background import BackgroundScheduler
import humanize
from flask_wtf.csrf import CSRFProtect, CSRFError
from openpyxl import Workbook
from io import BytesIO

app = Flask(__name__)
app.config['SECRET_KEY'] = 'ваш_секретний_ключ'  # Переконайтеся, що тут стоїть безпечний ключ
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///inventory.db'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # Обмеження розміру файлу (16 МБ)
app.config['ALLOWED_EXTENSIONS'] = {'png', 'jpg', 'jpeg', 'gif'}
app.config['BACKUP_FOLDER'] = 'backups'

# Створення директорії для завантажень, якщо вона не існує
if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

# Створення директорії для резервних копій, якщо вона не існує
if not os.path.exists(app.config['BACKUP_FOLDER']):
    os.makedirs(app.config['BACKUP_FOLDER'])

db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
csrf = CSRFProtect(app)

# Додаємо фільтр для перетворення переносів рядків у HTML-теги <br>
@app.template_filter('nl2br')
def nl2br(value):
    if value:
        return re.sub(r'\n', '<br>', value)
    return ''

# Моделі даних
class City(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False, unique=True)
    devices = db.relationship('Device', backref='city', lazy=True)
    
    def __repr__(self):
        return self.name

class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password = db.Column(db.String(120), nullable=False)
    is_admin = db.Column(db.Boolean, default=False)
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    city_id = db.Column(db.Integer, db.ForeignKey('city.id'))
    city = db.relationship('City', backref='users')

    def get_id(self):
        return str(self.id)

class Device(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    type = db.Column(db.String(50), nullable=False)
    serial_number = db.Column(db.String(100), unique=True)
    inventory_number = db.Column(db.String(20), unique=True)
    location = db.Column(db.String(200))
    status = db.Column(db.String(50))
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    city_id = db.Column(db.Integer, db.ForeignKey('city.id'), nullable=False)
    photos = db.relationship('DevicePhoto', backref='device', lazy=True, cascade="all, delete-orphan")
    history = db.relationship('DeviceHistory', backref='device', lazy=True, cascade="all, delete-orphan")
    last_maintenance = db.Column(db.DateTime, nullable=True)
    maintenance_interval = db.Column(db.Integer, default=365, nullable=True)
    
    def update_next_maintenance(self):
        """Оновлює дату наступного обслуговування на основі last_maintenance та maintenance_interval"""
        if self.last_maintenance and self.maintenance_interval:
            self.next_maintenance = self.last_maintenance + timedelta(days=self.maintenance_interval)
        else:
            self.next_maintenance = None
            
    next_maintenance = db.Column(db.DateTime, nullable=True)

class DevicePhoto(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    filename = db.Column(db.String(255), nullable=False)
    original_filename = db.Column(db.String(255), nullable=False)
    device_id = db.Column(db.Integer, db.ForeignKey('device.id'), nullable=False)
    uploaded_at = db.Column(db.DateTime, default=datetime.utcnow)

class DeviceHistory(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    device_id = db.Column(db.Integer, db.ForeignKey('device.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    action = db.Column(db.String(50), nullable=False)  # 'create', 'update', 'delete'
    field = db.Column(db.String(50))  # Яке поле було змінено
    old_value = db.Column(db.Text)  # Старе значення
    new_value = db.Column(db.Text)  # Нове значення
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)
    
    user = db.relationship('User', backref='device_actions')

class UserActivity(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    action = db.Column(db.String(100), nullable=False)
    ip_address = db.Column(db.String(45))  # IPv6 може бути довшим
    url = db.Column(db.String(255))
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)
    
    user = db.relationship('User', backref='activities')

class BackupSettings(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    enable_auto_backup = db.Column(db.Boolean, default=True)
    backup_interval = db.Column(db.String(20), default='daily')  # 'daily', 'weekly', 'monthly'
    retention_period = db.Column(db.Integer, default=30)  # скільки днів зберігати копії
    last_backup_at = db.Column(db.DateTime)
    next_backup_at = db.Column(db.DateTime)
    
    @classmethod
    def get_settings(cls):
        """Отримати або створити налаштування резервного копіювання"""
        settings = cls.query.first()
        if not settings:
            settings = cls()
            db.session.add(settings)
            db.session.commit()
        return settings
    
    def update_next_backup(self):
        """Оновити дату наступного резервного копіювання"""
        now = datetime.utcnow()
        self.last_backup_at = now
        
        if self.backup_interval == 'daily':
            self.next_backup_at = now + timedelta(days=1)
        elif self.backup_interval == 'weekly':
            self.next_backup_at = now + timedelta(weeks=1)
        elif self.backup_interval == 'monthly':
            # Приблизно один місяць (30 днів)
            self.next_backup_at = now + timedelta(days=30)
        
        db.session.commit()

class Backup(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    filename = db.Column(db.String(255), nullable=False)
    file_path = db.Column(db.String(512), nullable=False)
    size_bytes = db.Column(db.Integer, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    is_auto = db.Column(db.Boolean, default=False)
    
    @property
    def size(self):
        """Повертає розмір файлу в людиночитабельному форматі"""
        return humanize.naturalsize(self.size_bytes)

class SystemSettings(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    key = db.Column(db.String(100), nullable=False, unique=True)
    value = db.Column(db.Text)
    description = db.Column(db.String(255))
    
    @classmethod
    def get(cls, key, default=None):
        """Отримати значення налаштування за ключем"""
        setting = cls.query.filter_by(key=key).first()
        if setting:
            return setting.value
        return default
    
    @classmethod
    def set(cls, key, value, description=None):
        """Встановити значення налаштування"""
        setting = cls.query.filter_by(key=key).first()
        if setting:
            setting.value = value
            if description:
                setting.description = description
        else:
            setting = cls(key=key, value=value, description=description)
            db.session.add(setting)
        
        db.session.commit()
        return setting

@login_manager.user_loader
def load_user(user_id):
    user = User.query.get(int(user_id))
    if user and user.is_active:
        return user
    return None

# Захист адміністративних функцій
def admin_required(func):
    def wrapper(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin:
            abort(403)
        return func(*args, **kwargs)
    wrapper.__name__ = func.__name__
    return wrapper

# Перевірка допустимих розширень файлів
def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

# Генерація унікального інвентарного номера
def generate_inventory_number(city_prefix=""):
    # Генеруємо випадковий номер і перевіряємо його унікальність
    while True:
        # Якщо вказано префікс міста, використовуємо його
        if city_prefix:
            city_code = city_prefix[:3].upper()
        else:
            city_code = "INV"
        
        # Поточний рік
        year = datetime.now().strftime("%y")
        
        # Випадкова послідовність із 6 цифр
        random_digits = ''.join(random.choices(string.digits, k=6))
        
        # Створюємо інвентарний номер у форматі "XXX-YY-ZZZZZZ", де XXX - код міста, YY - рік, ZZZZZZ - випадковий номер
        inventory_number = f"{city_code}-{year}-{random_digits}"
        
        # Перевіряємо, чи такий номер вже існує
        if not Device.query.filter_by(inventory_number=inventory_number).first():
            return inventory_number

# Створення адміністратора при першому запуску
def create_admin():
    # Створюємо базове місто, якщо його ще немає
    default_city = City.query.filter_by(name='Головний Офіс').first()
    if not default_city:
        default_city = City(name='Головний Офіс')
        db.session.add(default_city)
        try:
            db.session.flush()  # Отримуємо ID міста без повного коміту
        except:
            # Якщо місто вже існує, отримуємо його знову
            db.session.rollback()
            default_city = City.query.filter_by(name='Головний Офіс').first()
    
    # Додаємо інші великі міста
    cities = ['Львів', 'Одеса', 'Київ', 'Луцьк', 'Чернівці']
    for city_name in cities:
        if not City.query.filter_by(name=city_name).first():
            city = City(name=city_name)
            db.session.add(city)
    
    # Створюємо адміністратора
    admin = User.query.filter_by(username='admin').first()
    if not admin:
        admin = User(
            username='admin',
            password=generate_password_hash('admin123'),
            is_admin=True,
            city_id=default_city.id
        )
        db.session.add(admin)
    
    try:
        db.session.commit()
    except:
        db.session.rollback()

# Функція для запису історії змін
def record_device_history(device_id, user_id, action, field=None, old_value=None, new_value=None):
    # Переконуємось, що device_id не є None
    if device_id is None:
        app.logger.error(f"Спроба створити запис історії з NULL device_id: action={action}, user_id={user_id}")
        return
        
    history = DeviceHistory(
        device_id=device_id,
        user_id=user_id,
        action=action,
        field=field,
        old_value=str(old_value) if old_value is not None else None,
        new_value=str(new_value) if new_value is not None else None
    )
    db.session.add(history)

# Функція для логування активності користувачів
def log_user_activity(user_id, action, ip_address=None, url=None):
    activity = UserActivity(
        user_id=user_id,
        action=action,
        ip_address=ip_address,
        url=url
    )
    db.session.add(activity)
    db.session.commit()

# Декоратор для логування дій користувача
def log_activity(action_description):
    def decorator(f):
        def wrapper(*args, **kwargs):
            # Виконуємо оригінальну функцію
            result = f(*args, **kwargs)
            
            # Логуємо активність, якщо користувач автентифікований
            if current_user.is_authenticated:
                log_user_activity(
                    current_user.id, 
                    action_description,
                    request.remote_addr,
                    request.url
                )
                
            return result
        wrapper.__name__ = f.__name__
        return wrapper
    return decorator

# Маршрути
@app.route('/')
def index():
    return render_template('index.html')



@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        user = User.query.filter_by(username=username).first()
        
        if user and user.is_active and check_password_hash(user.password, password):
            login_user(user)
            log_user_activity(user.id, 'Вхід до системи', request.remote_addr, request.url)
            return redirect(url_for('index'))
        else:
            return render_template('login.html', error='Невірний логін або пароль, або обліковий запис заблоковано')
            
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    log_user_activity(current_user.id, 'Вихід із системи', request.remote_addr, request.url)
    logout_user()
    return redirect(url_for('index'))

@app.route('/devices')
@login_required
def devices():
    selected_city_id = request.args.get('city_id', type=int)
    page = request.args.get('page', 1, type=int)
    per_page = 10  # кількість записів на сторінку
    
    # Для адміністраторів - показувати всі пристрої або з вибраного міста
    if current_user.is_admin:
        cities = City.query.all()
        if selected_city_id:
            query = Device.query.filter_by(city_id=selected_city_id)
        else:
            query = Device.query
    # Для звичайних користувачів - тільки пристрої їхнього міста
    else:
        cities = [current_user.city]
        query = Device.query.filter_by(city_id=current_user.city_id)
        selected_city_id = current_user.city_id
    
    # Застосовуємо пагінацію
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    devices = pagination.items
    
    return render_template('devices.html', devices=devices, cities=cities, 
                          selected_city_id=selected_city_id, pagination=pagination)

@app.route('/device/add', methods=['GET', 'POST'])
@login_required
@log_activity('Перегляд форми додавання пристрою')
def add_device():
    if request.method == 'POST':
        # Перевіряємо чи існує пристрій з таким серійним номером
        serial_number = request.form['serial_number']
        existing_device = Device.query.filter_by(serial_number=serial_number).first()
        if existing_device:
            flash('Пристрій з таким серійним номером вже існує!', 'error')
            if current_user.is_admin:
                cities = City.query.all()
            else:
                cities = [current_user.city]
            return render_template('add_device.html', cities=cities)

        # Визначаємо місто для пристрою
        if current_user.is_admin and request.form.get('city_id'):
            city_id = request.form.get('city_id', type=int)
            city = City.query.get(city_id)
            city_prefix = city.name if city else ""
        else:
            city_id = current_user.city_id
            city_prefix = current_user.city.name if current_user.city else ""
        
        # Генеруємо унікальний інвентарний номер
        inventory_number = generate_inventory_number(city_prefix)
        
        # Отримуємо дату останнього обслуговування
        last_maintenance = request.form.get('last_maintenance')
        maintenance_interval = request.form.get('maintenance_interval', 365, type=int)
        
        if last_maintenance:
            try:
                last_maintenance_date = datetime.strptime(last_maintenance, '%Y-%m-%d')
            except ValueError:
                last_maintenance_date = None
        else:
            last_maintenance_date = None
            
        device = Device(
            name=request.form['name'],
            type=request.form['type'],
            serial_number=request.form['serial_number'],
            inventory_number=inventory_number,
            location=request.form['location'],
            status=request.form['status'],
            notes=request.form['notes'],
            city_id=city_id,
            last_maintenance=last_maintenance_date,
            maintenance_interval=maintenance_interval
        )
        
        # Оновлюємо дату наступного обслуговування
        device.update_next_maintenance()
        
        db.session.add(device)
        db.session.flush()  # Отримуємо ID без коміту
        
        # Записуємо історію створення
        record_device_history(device.id, current_user.id, 'create')
        
        # Обробка файлів, якщо вони були завантажені
        if 'photos' in request.files:
            photos = request.files.getlist('photos')
            for photo in photos:
                if photo and allowed_file(photo.filename):
                    # Генеруємо унікальне ім'я файлу
                    filename = secure_filename(photo.filename)
                    unique_filename = f"{uuid.uuid4()}_{filename}"
                    
                    # Зберігаємо файл
                    photo.save(os.path.join(app.config['UPLOAD_FOLDER'], unique_filename))
                    
                    # Створюємо запис у базі даних
                    device_photo = DevicePhoto(
                        filename=unique_filename,
                        original_filename=filename,
                        device_id=device.id
                    )
                    db.session.add(device_photo)
        
        db.session.commit()
        log_user_activity(current_user.id, f'Додано новий пристрій: {device.name}', request.remote_addr, request.url)
        flash('Пристрій успішно додано!')
        return redirect(url_for('devices'))
    
    # Отримуємо список міст для адміністраторів
    if current_user.is_admin:
        cities = City.query.all()
    else:
        cities = [current_user.city]
        
    return render_template('add_device.html', cities=cities)

@app.route('/device/<int:device_id>')
@login_required
def device_detail(device_id):
    device = Device.query.get_or_404(device_id)
    
    # Перевіряємо, чи має користувач доступ до цього пристрою
    if not current_user.is_admin and device.city_id != current_user.city_id:
        abort(403)
        
    return render_template('device_detail.html', device=device, now=datetime.utcnow())

@app.route('/device/<int:device_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_device(device_id):
    device = Device.query.get_or_404(device_id)
    
    # Перевіряємо, чи має користувач доступ до цього пристрою
    if not current_user.is_admin and device.city_id != current_user.city_id:
        abort(403)
    
    if request.method == 'POST':
        # Визначаємо місто для пристрою
        if current_user.is_admin and request.form.get('city_id'):
            new_city_id = request.form.get('city_id', type=int)
            if device.city_id != new_city_id:
                old_city = device.city.name if device.city else "Не вказано"
                new_city = City.query.get(new_city_id).name
                record_device_history(device.id, current_user.id, 'update', 'city', old_city, new_city)
                device.city_id = new_city_id
        
        # Відстежуємо зміни основних полів
        fields_to_track = {
            'name': 'Назва',
            'type': 'Тип',
            'serial_number': 'Серійний номер',
            'location': 'Розташування',
            'status': 'Статус',
            'notes': 'Примітки'
        }
        
        for field, label in fields_to_track.items():
            old_value = getattr(device, field)
            new_value = request.form[field]
            if old_value != new_value:
                record_device_history(device.id, current_user.id, 'update', label, old_value, new_value)
                setattr(device, field, new_value)
        
        # Обробляємо дані обслуговування
        last_maintenance = request.form.get('last_maintenance')
        if last_maintenance:
            try:
                last_maintenance_date = datetime.strptime(last_maintenance, '%Y-%m-%d')
                if device.last_maintenance != last_maintenance_date:
                    old_value = device.last_maintenance.strftime('%Y-%m-%d') if device.last_maintenance else "Не вказано"
                    record_device_history(device.id, current_user.id, 'update', 'Дата обслуговування', old_value, last_maintenance)
                    device.last_maintenance = last_maintenance_date
            except ValueError:
                pass  # Ігноруємо неправильний формат дати
        else:
            if device.last_maintenance:
                old_value = device.last_maintenance.strftime('%Y-%m-%d')
                record_device_history(device.id, current_user.id, 'update', 'Дата обслуговування', old_value, "Не вказано")
                device.last_maintenance = None
        
        maintenance_interval = request.form.get('maintenance_interval', type=int)
        if maintenance_interval and device.maintenance_interval != maintenance_interval:
            old_value = str(device.maintenance_interval) if device.maintenance_interval else "365"
            record_device_history(device.id, current_user.id, 'update', 'Інтервал обслуговування', old_value, str(maintenance_interval))
            device.maintenance_interval = maintenance_interval
        
        # Оновлюємо дату наступного обслуговування
        device.update_next_maintenance()
        
        db.session.commit()
        flash('Пристрій успішно оновлено!')
        return redirect(url_for('device_detail', device_id=device.id))
    
    # Отримуємо список міст для адміністраторів
    if current_user.is_admin:
        cities = City.query.all()
    else:
        cities = [current_user.city]
        
    return render_template('edit_device.html', device=device, cities=cities)

@app.route('/device/<int:device_id>/delete', methods=['POST'])
@login_required
def delete_device(device_id):
    device = Device.query.get_or_404(device_id)
    
    # Перевіряємо, чи має користувач доступ до цього пристрою
    if not current_user.is_admin and device.city_id != current_user.city_id:
        abort(403)
    
    # Записуємо історію видалення
    record_device_history(device.id, current_user.id, 'delete')
    
    # Видаляємо фотографії пристрою з файлової системи
    for photo in device.photos:
        try:
            os.remove(os.path.join(app.config['UPLOAD_FOLDER'], photo.filename))
        except:
            # Якщо файл не знайдено, просто продовжуємо
            pass
    
    # Видаляємо пристрій (каскадне видалення також видалить записи фото)
    db.session.delete(device)
    db.session.commit()
    
    flash('Пристрій успішно видалено!')
    return redirect(url_for('devices'))

@app.route('/device/<int:device_id>/print')
@login_required
def print_inventory(device_id):
    device = Device.query.get_or_404(device_id)
    
    # Перевіряємо, чи має користувач доступ до цього пристрою
    if not current_user.is_admin and device.city_id != current_user.city_id:
        abort(403)
        
    return render_template('print_inventory.html', device=device)

@app.route('/devices/bulk-print', methods=['GET', 'POST'])
@login_required
def bulk_print_inventory():
    selected_city_id = request.args.get('city_id', type=int)
    
    # Для адміністраторів - показувати всі пристрої або з вибраного міста
    if current_user.is_admin:
        cities = City.query.all()
        if selected_city_id:
            devices = Device.query.filter_by(city_id=selected_city_id).all()
        else:
            devices = Device.query.all()
    # Для звичайних користувачів - тільки пристрої їхнього міста
    else:
        cities = [current_user.city]
        devices = Device.query.filter_by(city_id=current_user.city_id).all()
        selected_city_id = current_user.city_id
    
    # Якщо масовий друк був відправлений
    if request.method == 'POST':
        device_ids = request.form.getlist('device_ids', type=int)
        selected_devices = []
        
        for device_id in device_ids:
            device = Device.query.get_or_404(device_id)
            # Перевіряємо, чи має користувач доступ до цього пристрою
            if current_user.is_admin or device.city_id == current_user.city_id:
                selected_devices.append(device)
        
        if selected_devices:
            return render_template('bulk_print_inventory.html', devices=selected_devices)
        else:
            flash('Не вибрано жодного пристрою для друку!', 'warning')
    
    return render_template('bulk_print_select.html', devices=devices, cities=cities, selected_city_id=selected_city_id)

@app.route('/devices/export-excel')
@login_required
def export_excel():
    selected_city_id = request.args.get('city_id', type=int)
    
    # Для адміністраторів - показувати всі пристрої або з вибраного міста
    if current_user.is_admin:
        if selected_city_id:
            devices = Device.query.filter_by(city_id=selected_city_id).all()
            city_name = City.query.get(selected_city_id).name if selected_city_id else "Всі міста"
        else:
            devices = Device.query.all()
            city_name = "Всі міста"
    # Для звичайних користувачів - тільки пристрої їхнього міста
    else:
        devices = Device.query.filter_by(city_id=current_user.city_id).all()
        city_name = current_user.city.name
    
    # Створюємо новий Excel-файл
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Інвентаризація"
    
    # Стилі
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )
    
    # Заголовки
    headers = [
        "ID", "Інв. номер", "Назва", "Тип", "Серійний номер", 
        "Місто", "Розташування", "Статус", "Останнє обслуговування", 
        "Примітки", "Дата створення"
    ]
    
    # Ширина стовпців
    column_widths = {
        'A': 5,   # ID
        'B': 20,  # Інв. номер
        'C': 25,  # Назва
        'D': 15,  # Тип
        'E': 15,  # Серійний номер
        'F': 15,  # Місто
        'G': 20,  # Розташування
        'H': 15,  # Статус
        'I': 20,  # Останнє обслуговування
        'J': 30,  # Примітки
        'K': 20,  # Дата створення
    }
    
    # Встановлюємо ширину стовпців
    for col, width in column_widths.items():
        worksheet.column_dimensions[col].width = width
    
    # Додаємо заголовки
    for col_idx, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Додаємо дані
    for row_idx, device in enumerate(devices, 2):
        row_data = [
            device.id,
            device.inventory_number,
            device.name,
            device.type,
            device.serial_number,
            device.city.name,
            device.location,
            device.status,
            device.last_maintenance.strftime('%d.%m.%Y') if device.last_maintenance else "Не вказано",
            device.notes,
            device.created_at.strftime('%d.%m.%Y %H:%M')
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
    
    # Додаємо фільтри до заголовків
    worksheet.auto_filter.ref = f"A1:K{len(devices) + 1}"
    
    # Заморожуємо верхній рядок
    worksheet.freeze_panes = "A2"
    
    # Формуємо ім'я файлу
    current_date = datetime.now().strftime('%Y-%m-%d')
    filename = f"Інвентаризація_{city_name}_{current_date}.xlsx"
    
    # Зберігаємо файл у пам'яті
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    
    # Відправляємо файл користувачу
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )

@app.route('/device/<int:device_id>/add_photo', methods=['POST'])
@login_required
def add_device_photo(device_id):
    device = Device.query.get_or_404(device_id)
    
    # Перевіряємо, чи має користувач доступ до цього пристрою
    if not current_user.is_admin and device.city_id != current_user.city_id:
        abort(403)
    
    if 'photo' in request.files:
        photo = request.files['photo']
        if photo and allowed_file(photo.filename):
            # Генеруємо унікальне ім'я файлу
            filename = secure_filename(photo.filename)
            unique_filename = f"{uuid.uuid4()}_{filename}"
            
            # Зберігаємо файл
            photo.save(os.path.join(app.config['UPLOAD_FOLDER'], unique_filename))
            
            # Створюємо запис у базі даних
            device_photo = DevicePhoto(
                filename=unique_filename,
                original_filename=filename,
                device_id=device.id
            )
            db.session.add(device_photo)
            db.session.commit()
            
            flash('Фото успішно додано!')
    
    return redirect(url_for('device_detail', device_id=device_id))

@app.route('/device/photo/<int:photo_id>/delete', methods=['POST'])
@login_required
def delete_device_photo(photo_id):
    photo = DevicePhoto.query.get_or_404(photo_id)
    device = photo.device
    
    # Перевіряємо, чи має користувач доступ до цього пристрою
    if not current_user.is_admin and device.city_id != current_user.city_id:
        abort(403)
    
    # Видаляємо файл
    try:
        os.remove(os.path.join(app.config['UPLOAD_FOLDER'], photo.filename))
    except:
        # Якщо файл не знайдено, просто продовжуємо
        pass
    
    # Видаляємо запис з бази даних
    db.session.delete(photo)
    db.session.commit()
    
    flash('Фото успішно видалено!')
    return redirect(url_for('device_detail', device_id=device.id))

@app.route('/uploads/<filename>')
@login_required
def uploaded_file(filename):
    # Отримуємо фото, щоб перевірити права доступу
    photo = DevicePhoto.query.filter_by(filename=filename).first_or_404()
    device = photo.device
    
    # Перевіряємо, чи має користувач доступ до цього пристрою
    if not current_user.is_admin and device.city_id != current_user.city_id:
        abort(403)
        
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

# Адміністрування користувачів
@app.route('/admin/users')
@login_required
@admin_required
def admin_users():
    users = User.query.all()
    return render_template('admin/users.html', users=users)

@app.route('/admin/user/add', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_add_user():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        is_admin = 'is_admin' in request.form
        city_id = request.form.get('city_id', type=int)
        
        # Перевірка чи користувач вже існує
        existing_user = User.query.filter_by(username=username).first()
        if existing_user:
            flash('Користувач з таким іменем вже існує!', 'danger')
            cities = City.query.all()
            return render_template('admin/add_user.html', cities=cities)
            
        new_user = User(
            username=username,
            password=generate_password_hash(password),
            is_admin=is_admin,
            city_id=city_id
        )
        db.session.add(new_user)
        db.session.commit()
        flash('Користувача успішно додано!')
        return redirect(url_for('admin_users'))
    
    cities = City.query.all()
    return render_template('admin/add_user.html', cities=cities)

@app.route('/admin/user/edit/<int:user_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_edit_user(user_id):
    user = User.query.get_or_404(user_id)
    
    # Заборонити редагування власного облікового запису адміністратора
    if user.id == current_user.id:
        flash('Ви не можете редагувати власний обліковий запис через цю форму!', 'danger')
        return redirect(url_for('admin_users'))
    
    if request.method == 'POST':
        username = request.form['username']
        
        # Перевірка чи нове ім'я вже зайняте
        existing_user = User.query.filter(User.username == username, User.id != user_id).first()
        if existing_user:
            flash('Користувач з таким іменем вже існує!', 'danger')
            cities = City.query.all()
            return render_template('admin/edit_user.html', user=user, cities=cities)
            
        user.username = username
        
        # Оновлення паролю лише якщо вказано новий
        if request.form['password']:
            user.password = generate_password_hash(request.form['password'])
            
        user.is_admin = 'is_admin' in request.form
        user.is_active = 'is_active' in request.form
        user.city_id = request.form.get('city_id', type=int)
        
        db.session.commit()
        flash('Користувача успішно оновлено!')
        return redirect(url_for('admin_users'))
    
    cities = City.query.all()
    return render_template('admin/edit_user.html', user=user, cities=cities)

@app.route('/admin/user/toggle/<int:user_id>')
@login_required
@admin_required
def admin_toggle_user(user_id):
    user = User.query.get_or_404(user_id)
    
    # Заборонити блокування власного облікового запису
    if user.id == current_user.id:
        flash('Ви не можете заблокувати власний обліковий запис!', 'danger')
        return redirect(url_for('admin_users'))
    
    user.is_active = not user.is_active
    db.session.commit()
    
    action = "розблоковано" if user.is_active else "заблоковано"
    flash(f'Користувача {user.username} успішно {action}!')
    return redirect(url_for('admin_users'))

@app.route('/admin/user/delete/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def admin_delete_user(user_id):
    user = User.query.get_or_404(user_id)
    
    # Заборонити видалення власного облікового запису
    if user.id == current_user.id:
        flash('Ви не можете видалити власний обліковий запис!', 'danger')
        return redirect(url_for('admin_users'))
    
    db.session.delete(user)
    db.session.commit()
    flash(f'Користувача {user.username} успішно видалено!')
    return redirect(url_for('admin_users'))

# Адміністрування міст
@app.route('/admin/cities')
@login_required
@admin_required
def admin_cities():
    cities = City.query.all()
    return render_template('admin/cities.html', cities=cities)

@app.route('/admin/city/add', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_add_city():
    if request.method == 'POST':
        name = request.form['name']
        
        # Перевірка чи місто вже існує
        existing_city = City.query.filter_by(name=name).first()
        if existing_city:
            flash('Місто з такою назвою вже існує!', 'danger')
            return render_template('admin/add_city.html')
            
        new_city = City(name=name)
        db.session.add(new_city)
        db.session.commit()
        flash('Місто успішно додано!')
        return redirect(url_for('admin_cities'))
        
    return render_template('admin/add_city.html')

@app.route('/admin/city/edit/<int:city_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_edit_city(city_id):
    city = City.query.get_or_404(city_id)
    
    if request.method == 'POST':
        name = request.form['name']
        
        # Перевірка чи нова назва вже зайнята
        existing_city = City.query.filter(City.name == name, City.id != city_id).first()
        if existing_city:
            flash('Місто з такою назвою вже існує!', 'danger')
            return render_template('admin/edit_city.html', city=city)
            
        city.name = name
        db.session.commit()
        flash('Місто успішно оновлено!')
        return redirect(url_for('admin_cities'))
        
    return render_template('admin/edit_city.html', city=city)

@app.route('/admin/city/delete/<int:city_id>', methods=['POST'])
@login_required
@admin_required
def admin_delete_city(city_id):
    city = City.query.get_or_404(city_id)
    
    # Перевірка, чи є пристрої або користувачі, прив'язані до цього міста
    if Device.query.filter_by(city_id=city_id).first() or User.query.filter_by(city_id=city_id).first():
        flash('Неможливо видалити місто, оскільки існують пристрої або користувачі, прив\'язані до нього!', 'danger')
        return redirect(url_for('admin_cities'))
    
    db.session.delete(city)
    db.session.commit()
    flash(f'Місто {city.name} успішно видалено!')
    return redirect(url_for('admin_cities'))

@app.route('/admin/dashboard')
@login_required
@admin_required
def admin_dashboard():
    # Загальна статистика
    total_devices = Device.query.count()
    total_users = User.query.count()
    total_cities = City.query.count()
    
    # Розподіл пристроїв за містами
    devices_by_city = db.session.query(
        City.name, func.count(Device.id)
    ).join(City).group_by(City.name).all()
    
    # Розподіл пристроїв за типом
    devices_by_type = db.session.query(
        Device.type, func.count(Device.id)
    ).group_by(Device.type).all()
    
    # Розподіл пристроїв за статусом
    devices_by_status = db.session.query(
        Device.status, func.count(Device.id)
    ).group_by(Device.status).all()
    
    # Нові пристрої за останні 30 днів
    thirty_days_ago = datetime.utcnow() - timedelta(days=30)
    new_devices_count = Device.query.filter(Device.created_at >= thirty_days_ago).count()
    
    # Пристрої, додані за останні 12 місяців
    devices_by_month = []
    for i in range(12):
        start_date = datetime.utcnow().replace(day=1) - timedelta(days=30*i)
        end_date = (start_date.replace(day=28) + timedelta(days=4)).replace(day=1)
        count = Device.query.filter(Device.created_at >= start_date, Device.created_at < end_date).count()
        month_name = start_date.strftime('%B %Y')
        devices_by_month.append((month_name, count))
    devices_by_month.reverse()
    
    return render_template('admin/dashboard.html', 
                           total_devices=total_devices,
                           total_users=total_users,
                           total_cities=total_cities,
                           devices_by_city=devices_by_city,
                           devices_by_type=devices_by_type,
                           devices_by_status=devices_by_status,
                           new_devices_count=new_devices_count,
                           devices_by_month=devices_by_month)

@app.route('/device/<int:device_id>/history')
@login_required
def device_history(device_id):
    device = Device.query.get_or_404(device_id)
    
    # Перевіряємо, чи має користувач доступ до цього пристрою
    if not current_user.is_admin and device.city_id != current_user.city_id:
        abort(403)
    
    history = DeviceHistory.query.filter_by(device_id=device_id).order_by(DeviceHistory.timestamp.desc()).all()
    return render_template('device_history.html', device=device, history=history)

@app.route('/admin/user-activity')
@login_required
@admin_required
def admin_user_activity():
    page = request.args.get('page', 1, type=int)
    per_page = 50  # Збільшуємо кількість записів на сторінку для журналу
    
    activities = UserActivity.query.order_by(UserActivity.timestamp.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    return render_template('admin/user_activity.html', activities=activities)

@app.route('/toggle-theme')
def toggle_theme():
    theme = request.cookies.get('theme', 'light')
    new_theme = 'dark' if theme == 'light' else 'light'
    
    response = redirect(request.referrer or url_for('index'))
    response.set_cookie('theme', new_theme, max_age=60*60*24*365)  # Зберігаємо на рік
    
    return response

@app.context_processor
def utility_processor():
    def get_user_agent():
        return request.headers.get('User-Agent')
    
    return dict(get_user_agent=get_user_agent)

@app.route('/device/<int:device_id>/qrcode')
@login_required
def device_qrcode(device_id):
    device = Device.query.get_or_404(device_id)
    
    # Перевіряємо, чи має користувач доступ до цього пристрою
    if not current_user.is_admin and device.city_id != current_user.city_id:
        abort(403)
    
    # Створюємо QR-код з інформацією про пристрій
    qr_data = f"""
    ID: {device.id}
    Назва: {device.name}
    Інв. номер: {device.inventory_number}
    Тип: {device.type}
    S/N: {device.serial_number}
    Розташування: {device.location}
    Статус: {device.status}
    Місто: {device.city.name}
    """
    
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(qr_data)
    qr.make(fit=True)

    img = qr.make_image(fill_color="black", back_color="white")
    
    # Зберігаємо QR-код у буфер пам'яті
    img_buffer = io.BytesIO()
    img.save(img_buffer)
    img_buffer.seek(0)
    
    return send_file(
        img_buffer,
        mimetype='image/png',
        as_attachment=True,
        download_name=f'qrcode_{device.inventory_number}.png'
    )

@app.route('/device/<int:device_id>/print_qrcode')
@login_required
def print_device_qrcode(device_id):
    """Відображення сторінки для друку QR-коду пристрою"""
    device = Device.query.get_or_404(device_id)
    
    # Перевіряємо, чи має користувач доступ до цього пристрою
    if not current_user.is_admin and device.city_id != current_user.city_id:
        abort(403)
    
    return render_template('print_qrcode.html', device=device)

@app.route('/dashboard')
@login_required
def dashboard():
    return render_template('dashboard.html')

@app.route('/devices/import-excel', methods=['GET', 'POST'])
@login_required
@admin_required
def import_excel():
    if request.method == 'POST':
        if 'excel_file' not in request.files:
            flash('Файл не завантажено', 'danger')
            return redirect(request.url)
        
        file = request.files['excel_file']
        if file.filename == '':
            flash('Файл не вибрано', 'danger')
            return redirect(request.url)
        
        if file and file.filename.endswith(('.xlsx', '.xls')):
            try:
                # Зчитуємо дані з Excel
                import_wb = openpyxl.load_workbook(file)
                import_ws = import_wb.active
                
                # Перевіряємо заголовки
                headers = [cell.value for cell in import_ws[1]]
                required_headers = ['Назва', 'Тип', 'Серійний номер', 'Розташування', 'Статус', 'Місто', 'Примітки']
                
                # Перевіряємо, чи всі потрібні заголовки присутні
                missing_headers = [h for h in required_headers if h not in headers]
                if missing_headers:
                    flash(f'У файлі відсутні обов\'язкові стовпці: {", ".join(missing_headers)}', 'danger')
                    return redirect(request.url)
                
                # Індекси стовпців
                name_idx = headers.index('Назва')
                type_idx = headers.index('Тип')
                serial_idx = headers.index('Серійний номер')
                location_idx = headers.index('Розташування')
                status_idx = headers.index('Статус')
                city_idx = headers.index('Місто')
                notes_idx = headers.index('Примітки')
                
                # Читаємо дані, починаючи з другого рядка
                success_count = 0
                error_count = 0
                errors = []
                
                for row_idx, row in enumerate(import_ws.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        name = row[name_idx]
                        device_type = row[type_idx]
                        serial_number = row[serial_idx]
                        location = row[location_idx]
                        status = row[status_idx]
                        city_name = row[city_idx]
                        notes = row[notes_idx] if row[notes_idx] else ""
                        
                        # Пропускаємо, якщо основні поля відсутні
                        if not name or not device_type or not serial_number or not city_name:
                            errors.append(f'Рядок {row_idx}: Відсутні обов\'язкові дані')
                            error_count += 1
                            continue
                        
                        # Знаходимо або створюємо місто
                        city = City.query.filter_by(name=city_name).first()
                        if not city:
                            city = City(name=city_name)
                            db.session.add(city)
                            db.session.flush()
                        
                        # Перевіряємо, чи пристрій з таким серійним номером вже існує
                        existing_device = Device.query.filter_by(serial_number=serial_number).first()
                        if existing_device:
                            errors.append(f'Рядок {row_idx}: Пристрій з серійним номером {serial_number} вже існує')
                            error_count += 1
                            continue
                        
                        # Створюємо інвентарний номер
                        inventory_number = generate_inventory_number(city.name)
                        
                        # Створюємо новий пристрій
                        device = Device(
                            name=name,
                            type=device_type,
                            serial_number=serial_number,
                            inventory_number=inventory_number,
                            location=location,
                            status=status,
                            notes=notes,
                            city_id=city.id
                        )
                        
                        db.session.add(device)
                        db.session.flush()
                        
                        # Записуємо історію
                        record_device_history(device.id, current_user.id, 'create', 'Імпорт з Excel')
                        
                        success_count += 1
                    
                    except Exception as e:
                        errors.append(f'Рядок {row_idx}: {str(e)}')
                        error_count += 1
                
                # Зберігаємо зміни
                db.session.commit()
                
                # Формуємо повідомлення
                message = f'Імпортовано успішно: {success_count} пристроїв.'
                if error_count > 0:
                    message += f' З помилками: {error_count}.'
                    for error in errors[:10]:  # Показуємо перші 10 помилок
                        message += f'<br>{error}'
                    if len(errors) > 10:
                        message += f'<br>...і ще {len(errors) - 10} помилок.'
                
                if success_count > 0:
                    flash(message, 'success')
                else:
                    flash(message, 'warning')
                
                return redirect(url_for('devices'))
            
            except Exception as e:
                flash(f'Помилка при імпорті: {str(e)}', 'danger')
                return redirect(request.url)
        else:
            flash('Дозволено тільки файли Excel (.xlsx, .xls)', 'danger')
            return redirect(request.url)
    
    return render_template('import_excel.html')

@app.route('/template/download/device-import')
@login_required
@admin_required
def download_import_template():
    # Створюємо новий Excel-файл
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Шаблон імпорту"
    
    # Заголовки
    headers = ['Назва', 'Тип', 'Серійний номер', 'Розташування', 'Статус', 'Місто', 'Примітки']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    
    # Приклад даних
    example_data = [
        'Ноутбук Dell Latitude 5420', 'Ноутбук', 'SN12345678', 'Кабінет 205', 'В експлуатації', 'Київ', 'Новий ноутбук для бухгалтерії'
    ]
    for col_idx, value in enumerate(example_data, start=1):
        ws.cell(row=2, column=col_idx).value = value
    
    # Приклад даних 2
    example_data_2 = [
        'Принтер HP LaserJet Pro', 'Принтер', 'SNPR123456', 'Кабінет 105', 'В експлуатації', 'Львів', 'Мережевий принтер'
    ]
    for col_idx, value in enumerate(example_data_2, start=1):
        ws.cell(row=3, column=col_idx).value = value
    
    # Ширина стовпців
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20
    
    # Зберігаємо в пам'яті
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"template_device_import_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# Функції для роботи з резервними копіями

def create_backup(is_auto=False):
    """Створює резервну копію бази даних та завантажених файлів"""
    try:
        # Створюємо унікальне ім'я для резервної копії
        timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
        backup_type = 'auto' if is_auto else 'manual'
        backup_name = f"backup_{backup_type}_{timestamp}"
        backup_path = os.path.join(app.config['BACKUP_FOLDER'], backup_name)
        
        # Створюємо теку для резервної копії
        os.makedirs(backup_path, exist_ok=True)
        
        # Копіюємо базу даних
        db_path = os.path.join(app.instance_path, 'inventory.db')
        db_backup_path = os.path.join(backup_path, 'inventory.db')
        shutil.copy2(db_path, db_backup_path)
        
        # Копіюємо завантажені файли
        uploads_backup_path = os.path.join(backup_path, 'uploads')
        if os.path.exists(app.config['UPLOAD_FOLDER']):
            shutil.copytree(app.config['UPLOAD_FOLDER'], uploads_backup_path)
        
        # Створюємо ZIP-архів
        zip_filename = f"{backup_name}.zip"
        zip_path = os.path.join(app.config['BACKUP_FOLDER'], zip_filename)
        
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            # Додаємо базу даних
            zipf.write(db_backup_path, arcname='inventory.db')
            
            # Додаємо файли завантажень
            if os.path.exists(uploads_backup_path):
                for root, dirs, files in os.walk(uploads_backup_path):
                    for file in files:
                        file_path = os.path.join(root, file)
                        arcname = os.path.relpath(file_path, backup_path)
                        zipf.write(file_path, arcname=arcname)
        
        # Видаляємо тимчасову теку
        shutil.rmtree(backup_path)
        
        # Зберігаємо інформацію про резервну копію в базі даних
        file_size = os.path.getsize(zip_path)
        backup = Backup(
            filename=zip_filename,
            file_path=zip_path,
            size_bytes=file_size,
            is_auto=is_auto
        )
        db.session.add(backup)
        
        # Оновлюємо налаштування резервного копіювання
        if is_auto:
            backup_settings = BackupSettings.get_settings()
            backup_settings.update_next_backup()
        
        db.session.commit()
        
        # Видаляємо старі резервні копії
        cleanup_old_backups()
        
        return backup
    except Exception as e:
        # Логуємо помилку
        app.logger.error(f"Помилка створення резервної копії: {str(e)}")
        if os.path.exists(backup_path):
            shutil.rmtree(backup_path)
        return None

def restore_from_backup(backup_id=None, file_path=None):
    """Відновлює систему з резервної копії"""
    try:
        if backup_id:
            backup = Backup.query.get(backup_id)
            if not backup:
                return False, "Резервну копію не знайдено"
            file_path = backup.file_path
        
        if not file_path or not os.path.exists(file_path):
            return False, "Файл резервної копії не знайдено"
        
        # Створюємо тимчасову теку для розпакування
        temp_dir = os.path.join(app.config['BACKUP_FOLDER'], 'temp_restore')
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
        os.makedirs(temp_dir, exist_ok=True)
        
        # Розпаковуємо архів
        with zipfile.ZipFile(file_path, 'r') as zipf:
            zipf.extractall(temp_dir)
        
        # Закриваємо з'єднання з базою даних перед заміною файлу
        db.session.close()
        
        # Замінюємо файл бази даних
        db_path = os.path.join(app.instance_path, 'inventory.db')
        db_backup_path = os.path.join(temp_dir, 'inventory.db')
        
        # Створюємо резервну копію поточної бази перед заміною
        current_db_backup = os.path.join(app.config['BACKUP_FOLDER'], f"pre_restore_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.db")
        shutil.copy2(db_path, current_db_backup)
        
        # Замінюємо базу даних
        shutil.copy2(db_backup_path, db_path)
        
        # Відновлюємо завантажені файли, якщо вони є в архіві
        uploads_backup_path = os.path.join(temp_dir, 'uploads')
        if os.path.exists(uploads_backup_path):
            # Видаляємо поточну теку завантажень
            if os.path.exists(app.config['UPLOAD_FOLDER']):
                shutil.rmtree(app.config['UPLOAD_FOLDER'])
            
            # Копіюємо теку з резервної копії
            shutil.copytree(uploads_backup_path, app.config['UPLOAD_FOLDER'])
        
        # Видаляємо тимчасову теку
        shutil.rmtree(temp_dir)
        
        return True, "Система успішно відновлена з резервної копії"
    except Exception as e:
        app.logger.error(f"Помилка відновлення з резервної копії: {str(e)}")
        return False, f"Помилка відновлення: {str(e)}"

def cleanup_old_backups():
    """Видаляє старі резервні копії відповідно до налаштувань збереження"""
    try:
        settings = BackupSettings.get_settings()
        retention_days = settings.retention_period
        
        # Обчислюємо дату, до якої треба видалити старі копії
        cutoff_date = datetime.utcnow() - timedelta(days=retention_days)
        
        # Знаходимо старі копії
        old_backups = Backup.query.filter(Backup.created_at < cutoff_date).all()
        
        # Видаляємо їх
        for backup in old_backups:
            try:
                if os.path.exists(backup.file_path):
                    os.remove(backup.file_path)
                db.session.delete(backup)
            except Exception as e:
                app.logger.error(f"Помилка видалення старої резервної копії {backup.filename}: {str(e)}")
        
        db.session.commit()
    except Exception as e:
        app.logger.error(f"Помилка при очищенні старих резервних копій: {str(e)}")

def export_settings():
    """Експортує налаштування системи в JSON"""
    try:
        settings = SystemSettings.query.all()
        settings_dict = {setting.key: setting.value for setting in settings}
        
        # Додаємо налаштування резервного копіювання
        backup_settings = BackupSettings.get_settings()
        settings_dict['backup'] = {
            'enable_auto_backup': backup_settings.enable_auto_backup,
            'backup_interval': backup_settings.backup_interval,
            'retention_period': backup_settings.retention_period
        }
        
        return settings_dict
    except Exception as e:
        app.logger.error(f"Помилка експорту налаштувань: {str(e)}")
        return {}

def import_settings(settings_dict):
    """Імпортує налаштування системи з JSON"""
    try:
        # Імпортуємо загальні налаштування
        for key, value in settings_dict.items():
            if key != 'backup':  # Обробляємо налаштування резервного копіювання окремо
                SystemSettings.set(key, value)
        
        # Імпортуємо налаштування резервного копіювання, якщо вони є
        if 'backup' in settings_dict:
            backup_data = settings_dict['backup']
            backup_settings = BackupSettings.get_settings()
            
            if 'enable_auto_backup' in backup_data:
                backup_settings.enable_auto_backup = backup_data['enable_auto_backup']
            
            if 'backup_interval' in backup_data:
                backup_settings.backup_interval = backup_data['backup_interval']
            
            if 'retention_period' in backup_data:
                backup_settings.retention_period = backup_data['retention_period']
            
            backup_settings.update_next_backup()
            db.session.commit()
        
        return True, "Налаштування успішно імпортовано"
    except Exception as e:
        app.logger.error(f"Помилка імпорту налаштувань: {str(e)}")
        return False, f"Помилка імпорту налаштувань: {str(e)}"

# Функція для автоматичного резервного копіювання
def check_auto_backup():
    """Перевіряє, чи потрібно створити автоматичну резервну копію"""
    try:
        settings = BackupSettings.get_settings()
        
        # Перевіряємо, чи увімкнено автоматичне резервне копіювання
        if not settings.enable_auto_backup:
            return
        
        now = datetime.utcnow()
        
        # Якщо дата наступного резервного копіювання не встановлена
        # або ми вже за цією датою, створюємо резервну копію
        if not settings.next_backup_at or now >= settings.next_backup_at:
            create_backup(is_auto=True)
            # Оновлення дати наступного резервного копіювання
            # відбувається в функції create_backup
    except Exception as e:
        app.logger.error(f"Помилка при перевірці автоматичного резервного копіювання: {str(e)}")

# Ініціалізація планувальника для запуску автоматичного резервного копіювання
def init_scheduler():
    """Ініціалізує планувальник задач"""
    scheduler = BackgroundScheduler()
    scheduler.add_job(check_auto_backup, 'interval', hours=1)  # Перевіряємо щогодини
    scheduler.start()
    app.scheduler = scheduler  # Зберігаємо планувальник як атрибут додатку

# Додаємо перед іншими маршрутами
@app.route('/admin/backup')
@login_required
@admin_required
def admin_backup():
    """Сторінка управління резервними копіями"""
    backups = Backup.query.order_by(Backup.created_at.desc()).all()
    backup_settings = BackupSettings.get_settings()
    return render_template('admin/backup.html', backups=backups, backup_settings=backup_settings)

@app.route('/admin/backup/create', methods=['POST'])
@login_required
@admin_required
def admin_create_backup():
    """Створює резервну копію вручну"""
    backup = create_backup(is_auto=False)
    if backup:
        flash('Резервна копія успішно створена.')
        log_user_activity(current_user.id, 'Створення резервної копії', request.remote_addr, request.url)
    else:
        flash('Помилка при створенні резервної копії.', 'error')
    return redirect(url_for('admin_backup'))

@app.route('/admin/backup/settings', methods=['POST'])
@login_required
@admin_required
def admin_backup_settings():
    """Оновлює налаштування резервного копіювання"""
    backup_settings = BackupSettings.get_settings()
    
    # Отримуємо параметри з форми
    backup_settings.enable_auto_backup = 'enable_auto_backup' in request.form
    backup_settings.backup_interval = request.form.get('backup_interval', 'daily')
    try:
        backup_settings.retention_period = int(request.form.get('retention_period', 30))
    except ValueError:
        backup_settings.retention_period = 30
    
    # Оновлюємо дату наступного резервного копіювання
    backup_settings.update_next_backup()
    
    flash('Налаштування резервного копіювання успішно оновлені.')
    log_user_activity(current_user.id, 'Оновлення налаштувань резервного копіювання', request.remote_addr, request.url)
    return redirect(url_for('admin_backup'))

@app.route('/admin/backup/download/<int:backup_id>')
@login_required
@admin_required
def admin_download_backup(backup_id):
    """Завантажує резервну копію на комп'ютер користувача"""
    backup = Backup.query.get_or_404(backup_id)
    return send_file(backup.file_path, as_attachment=True, download_name=backup.filename)

@app.route('/admin/backup/restore/<int:backup_id>')
@login_required
@admin_required
def admin_restore_backup(backup_id):
    """Відновлює систему з резервної копії"""
    success, message = restore_from_backup(backup_id=backup_id)
    
    if success:
        flash('Система успішно відновлена з резервної копії.')
        log_user_activity(current_user.id, 'Відновлення системи з резервної копії', request.remote_addr, request.url)
    else:
        flash(f'Помилка при відновленні системи: {message}', 'error')
    
    return redirect(url_for('admin_backup'))

@app.route('/admin/backup/delete/<int:backup_id>')
@login_required
@admin_required
def admin_delete_backup(backup_id):
    """Видаляє резервну копію"""
    backup = Backup.query.get_or_404(backup_id)
    try:
        if os.path.exists(backup.file_path):
            os.remove(backup.file_path)
        db.session.delete(backup)
        db.session.commit()
        flash('Резервна копія успішно видалена.')
        log_user_activity(current_user.id, 'Видалення резервної копії', request.remote_addr, request.url)
    except Exception as e:
        flash(f'Помилка при видаленні резервної копії: {str(e)}', 'error')
    
    return redirect(url_for('admin_backup'))

@app.route('/admin/backup/restore-from-file', methods=['POST'])
@login_required
@admin_required
def admin_restore_from_file():
    """Відновлює систему з завантаженого файлу резервної копії"""
    if 'backup_file' not in request.files:
        flash('Файл резервної копії не було надано.', 'error')
        return redirect(url_for('admin_backup'))
    
    backup_file = request.files['backup_file']
    if backup_file.filename == '':
        flash('Не вибрано файл резервної копії.', 'error')
        return redirect(url_for('admin_backup'))
    
    # Зберігаємо завантажений файл
    temp_file_path = os.path.join(app.config['BACKUP_FOLDER'], f"temp_upload_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}.zip")
    backup_file.save(temp_file_path)
    
    # Відновлюємо систему
    success, message = restore_from_backup(file_path=temp_file_path)
    
    # Видаляємо тимчасовий файл
    if os.path.exists(temp_file_path):
        os.remove(temp_file_path)
    
    if success:
        flash('Система успішно відновлена з завантаженого файлу.')
        log_user_activity(current_user.id, 'Відновлення системи з завантаженого файлу', request.remote_addr, request.url)
    else:
        flash(f'Помилка при відновленні системи: {message}', 'error')
    
    return redirect(url_for('admin_backup'))

@app.route('/admin/settings/export')
@login_required
@admin_required
def admin_export_settings():
    """Експортує налаштування системи в JSON-файл"""
    settings_dict = export_settings()
    
    # Створюємо байтовий потік з JSON-даними
    json_data = json.dumps(settings_dict, indent=4, ensure_ascii=False).encode('utf-8')
    file_stream = io.BytesIO(json_data)
    
    # Формуємо ім'я файлу з датою експорту
    timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
    filename = f"settings_export_{timestamp}.json"
    
    log_user_activity(current_user.id, 'Експорт налаштувань системи', request.remote_addr, request.url)
    
    return send_file(
        file_stream,
        as_attachment=True,
        download_name=filename,
        mimetype='application/json'
    )

@app.route('/admin/settings/import', methods=['POST'])
@login_required
@admin_required
def admin_import_settings():
    # Перевіряємо, чи був завантажений файл
    if 'settings_file' not in request.files:
        flash('Файл налаштувань не завантажено', 'danger')
        return redirect(url_for('admin_backup'))
    
    file = request.files['settings_file']
    if file.filename == '':
        flash('Не вибрано файл', 'danger')
        return redirect(url_for('admin_backup'))
    
    if file and file.filename.endswith('.json'):
        try:
            # Читаємо файл налаштувань
            settings_data = json.loads(file.read().decode('utf-8'))
            
            # Імпортуємо налаштування
            import_settings(settings_data)
            
            flash('Налаштування успішно імпортовано', 'success')
            return redirect(url_for('admin_backup'))
        except Exception as e:
            app.logger.error(f"Помилка при імпорті налаштувань: {str(e)}")
            flash(f'Помилка імпорту налаштувань: {str(e)}', 'danger')
            return redirect(url_for('admin_backup'))
    else:
        flash('Дозволені тільки файли формату JSON', 'danger')
        return redirect(url_for('admin_backup'))

@app.errorhandler(400)
def handle_csrf_error(e):
    return render_template('error.html', error_code=400, error_message="Помилка запиту. Можливо, закінчився час сесії. Спробуйте оновити сторінку."), 400

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        create_admin()
        init_scheduler()  # Ініціалізуємо планувальник задач
    # app.run(host='krainamriy.fun', port=80, debug=True)  # Для продакшену
    app.run(debug=True)  # Для розробки