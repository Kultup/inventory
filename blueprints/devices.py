from flask import Blueprint, render_template, request, redirect, url_for, flash, abort, send_file, send_from_directory, current_app
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
from sqlalchemy import func, or_
from sqlalchemy.orm import joinedload, selectinload
import os
import uuid
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from datetime import datetime, date
import qrcode
from PIL import Image
import pytz

# Імпорти моделей та функцій
from models import Device, DevicePhoto, DeviceHistory, City, User, db, RepairExpense
from utils import (allowed_file, record_device_history, generate_inventory_number, log_user_activity, 
                   optimize_image, generate_thumbnails, convert_to_webp, cleanup_unused_photos)

devices_bp = Blueprint('devices', __name__)

# Rate limiting та кешування отримуються через current_app.extensions під час виконання

@devices_bp.route('/devices')
@login_required
def devices():
    # Параметри пагінації та фільтрації
    selected_city_id = request.args.get('city_id', type=int)
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)  # Збільшено до 20 записів
    
    # Параметри пошуку та фільтрації
    search = request.args.get('search', '').strip()
    device_type = request.args.get('type', '').strip()
    status = request.args.get('status', '').strip()
    sort_by = request.args.get('sort', 'created_at')
    sort_order = request.args.get('order', 'desc')
    
    # Розширені фільтри
    created_from = request.args.get('created_from', '').strip()
    created_to = request.args.get('created_to', '').strip()
    price_from = request.args.get('price_from', '').strip()
    price_to = request.args.get('price_to', '').strip()
    
    # Базовий запит з eager loading для city
    # Кешуємо список міст (TTL 1 година)
    cache_key = f'cities_{current_user.id if not current_user.is_admin else "all"}'
    try:
        cache_obj = current_app.extensions.get('cache')
        # Перевіряємо, чи це об'єкт Cache (має метод set)
        if cache_obj and hasattr(cache_obj, 'set'):
            cities = cache_obj.get(cache_key)
            if cities is None:
                if current_user.is_admin:
                    cities = City.query.all()
                else:
                    cities = [current_user.city]
                cache_obj.set(cache_key, cities, timeout=3600)  # 1 година
            else:
                # Дані з кешу
                pass
        else:
            # Кеш не доступний, отримуємо дані без кешування
            if current_user.is_admin:
                cities = City.query.all()
            else:
                cities = [current_user.city]
    except (KeyError, AttributeError, TypeError):
        # Якщо кеш не доступний, просто отримуємо дані без кешування
        if current_user.is_admin:
            cities = City.query.all()
        else:
            cities = [current_user.city]
    
    if current_user.is_admin:
        query = Device.query.options(joinedload(Device.city))
        if selected_city_id:
            query = query.filter_by(city_id=selected_city_id)
    else:
        query = Device.query.options(joinedload(Device.city)).filter_by(city_id=current_user.city_id)
        selected_city_id = current_user.city_id
    
    # Розширений пошук по всіх полях
    if search:
        search_filter = or_(
            Device.name.ilike(f'%{search}%'),
            Device.type.ilike(f'%{search}%'),
            Device.serial_number.ilike(f'%{search}%'),
            Device.inventory_number.ilike(f'%{search}%'),
            Device.location.ilike(f'%{search}%'),
            Device.notes.ilike(f'%{search}%')
        )
        query = query.filter(search_filter)
    
    # Фільтр по типу пристрою
    if device_type:
        query = query.filter(Device.type.ilike(f'%{device_type}%'))
    
    # Фільтр по статусу
    if status:
        query = query.filter(Device.status.ilike(f'%{status}%'))
    
    # Розширені фільтри
    if created_from:
        try:
            from datetime import datetime
            created_from_date = datetime.strptime(created_from, '%Y-%m-%d').date()
            query = query.filter(Device.created_at >= created_from_date)
        except ValueError:
            pass
    
    if created_to:
        try:
            from datetime import datetime
            created_to_date = datetime.strptime(created_to, '%Y-%m-%d').date()
            query = query.filter(Device.created_at <= created_to_date)
        except ValueError:
            pass
    
    if price_from:
        try:
            price_from_float = float(price_from)
            query = query.filter(Device.purchase_price >= price_from_float)
        except ValueError:
            pass
    
    if price_to:
        try:
            price_to_float = float(price_to)
            query = query.filter(Device.purchase_price <= price_to_float)
        except ValueError:
            pass
    
    # Сортування
    if sort_by in ['name', 'type', 'serial_number', 'inventory_number', 'location', 'status', 'created_at', 'last_maintenance']:
        sort_column = getattr(Device, sort_by)
        if sort_order == 'desc':
            query = query.order_by(sort_column.desc())
        else:
            query = query.order_by(sort_column.asc())
    else:
        query = query.order_by(Device.created_at.desc())
    
    # Застосовуємо пагінацію з оптимізацією
    pagination = query.paginate(
        page=page, 
        per_page=min(per_page, 100),  # Максимум 100 записів на сторінку
        error_out=False
    )
    devices = pagination.items
    
    # Отримуємо унікальні типи та статуси для фільтрів
    device_types = db.session.query(Device.type).distinct().filter(Device.type.isnot(None)).all()
    device_statuses = db.session.query(Device.status).distinct().filter(Device.status.isnot(None)).all()
    
    return render_template('devices.html', 
                          devices=devices, 
                          cities=cities,
                          selected_city_id=selected_city_id, 
                          pagination=pagination,
                          search=search,
                          device_type=device_type,
                          status=status,
                          sort_by=sort_by,
                          sort_order=sort_order,
                          device_types=[t[0] for t in device_types],
                          device_statuses=[s[0] for s in device_statuses],
                          per_page=per_page,
                          created_from=created_from,
                          created_to=created_to,
                          price_from=price_from,
                          price_to=price_to)

@devices_bp.route('/device/add', methods=['GET', 'POST'])
@login_required
def add_device():
    if request.method == 'POST':
        # Перевіряємо чи існує пристрій з таким серійним номером
        serial_number = request.form['serial_number']
        existing_device = Device.query.filter_by(serial_number=serial_number).first()
        if existing_device:
            flash('Пристрій з таким серійним номером вже існує!', 'error')
            if current_user.is_admin:
                cities = City.query.all()
            else:
                cities = [current_user.city]
            return render_template('add_device.html', cities=cities)

        # Визначаємо місто для пристрою
        if current_user.is_admin and request.form.get('city_id'):
            city_id = request.form.get('city_id', type=int)
            city = City.query.get(city_id)
            city_prefix = city.name if city else ""
        else:
            city_id = current_user.city_id
            city_prefix = current_user.city.name if current_user.city else ""
        
        # Генеруємо унікальний інвентарний номер
        inventory_number = generate_inventory_number()
        
        # Отримуємо дату останнього обслуговування
        last_maintenance = request.form.get('last_maintenance')
        maintenance_interval = request.form.get('maintenance_interval', 365, type=int)
        
        if last_maintenance:
            try:
                last_maintenance_date = datetime.strptime(last_maintenance, '%Y-%m-%d')
            except ValueError:
                last_maintenance_date = None
        else:
            last_maintenance_date = None
        
        # Отримуємо фінансову інформацію
        purchase_price = request.form.get('purchase_price')
        purchase_date = request.form.get('purchase_date')
        
        if purchase_price:
            try:
                purchase_price = float(purchase_price)
            except (ValueError, TypeError):
                purchase_price = None
        else:
            purchase_price = None
        
        if purchase_date:
            try:
                purchase_date = datetime.strptime(purchase_date, '%Y-%m-%d').date()
            except ValueError:
                purchase_date = None
        else:
            purchase_date = None
            
        device = Device(
            name=request.form['name'],
            type=request.form['type'],
            serial_number=request.form['serial_number'],
            inventory_number=inventory_number,
            location=request.form['location'],
            status=request.form['status'],
            notes=request.form['notes'],
            city_id=city_id,
            last_maintenance=last_maintenance_date,
            maintenance_interval=maintenance_interval,
            purchase_price=purchase_price,
            purchase_date=purchase_date
        )
        
        # Оновлюємо дату наступного обслуговування
        device.update_next_maintenance()
        
        db.session.add(device)
        db.session.flush()  # Отримуємо ID без коміту
        
        # Записуємо історію створення
        record_device_history(device.id, current_user.id, 'create')
        
        # Обробка файлів, якщо вони були завантажені
        if 'photos' in request.files:
            photos = request.files.getlist('photos')
            for photo in photos:
                if photo and allowed_file(photo.filename):
                    # Генеруємо унікальне ім'я файлу
                    filename = secure_filename(photo.filename)
                    unique_filename = f"{uuid.uuid4()}_{filename}"
                    
                    # Зберігаємо файл
                    file_path = os.path.join(current_app.config['UPLOAD_FOLDER'], unique_filename)
                    photo.save(file_path)
                    
                    # Оптимізуємо зображення
                    optimized_path = optimize_image(file_path)
                    if optimized_path and optimized_path != file_path:
                        # Якщо файл було конвертовано (наприклад PNG->JPG)
                        unique_filename = os.path.basename(optimized_path)
                        file_path = optimized_path
                    
                    # Генеруємо thumbnail'и
                    thumbnails = generate_thumbnails(file_path)
                    
                    # Створюємо WebP версію (якщо браузер підтримує)
                    webp_path = convert_to_webp(file_path)
                    
                    # Створюємо запис у базі даних
                    device_photo = DevicePhoto(
                        filename=unique_filename,
                        original_filename=filename,
                        device_id=device.id
                    )
                    db.session.add(device_photo)
        
        db.session.commit()
        log_user_activity(current_user.id, f'Додано новий пристрій: {device.name}', request.remote_addr, request.url)
        flash('Пристрій успішно додано!')
        return redirect(url_for('devices.devices'))
    
    # Отримуємо список міст для адміністраторів
    if current_user.is_admin:
        cities = City.query.all()
    else:
        cities = [current_user.city]
        
    return render_template('add_device.html', cities=cities)

@devices_bp.route('/device/<int:device_id>')
@login_required
def device_detail(device_id):
    device = Device.query.options(
        joinedload(Device.city),
        selectinload(Device.photos),
        selectinload(Device.repair_expenses)
    ).get_or_404(device_id)
    
    # Перевіряємо, чи має користувач доступ до цього пристрою
    if not current_user.is_admin and device.city_id != current_user.city_id:
        abort(403)
        
    return render_template('device_detail.html', device=device, now=date.today())

@devices_bp.route('/device/<int:device_id>/test-notification', methods=['POST'])
@login_required
def test_device_notification(device_id):
    abort(404)

@devices_bp.route('/device/<int:device_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_device(device_id):
    device = Device.query.get_or_404(device_id)
    
    # Перевіряємо, чи має користувач доступ до цього пристрою
    if not current_user.is_admin and device.city_id != current_user.city_id:
        abort(403)
    
    if request.method == 'POST':
        # Визначаємо місто для пристрою
        if current_user.is_admin and request.form.get('city_id'):
            new_city_id = request.form.get('city_id', type=int)
            if device.city_id != new_city_id:
                old_city = device.city.name if device.city else "Не вказано"
                new_city = City.query.get(new_city_id).name
                record_device_history(device.id, current_user.id, 'update', 'city', old_city, new_city)
                device.city_id = new_city_id
        
        # Відстежуємо зміни основних полів
        fields_to_track = {
            'name': 'Назва',
            'type': 'Тип',
            'serial_number': 'Серійний номер',
            'location': 'Розташування',
            'status': 'Статус',
            'notes': 'Примітки'
        }
        
        for field, label in fields_to_track.items():
            old_value = getattr(device, field)
            new_value = request.form[field]
            if old_value != new_value:
                record_device_history(device.id, current_user.id, 'update', label, old_value, new_value)
                setattr(device, field, new_value)
        
        # Обробляємо дані обслуговування
        last_maintenance = request.form.get('last_maintenance')
        if last_maintenance:
            try:
                last_maintenance_date = datetime.strptime(last_maintenance, '%Y-%m-%d')
                if device.last_maintenance != last_maintenance_date:
                    old_value = device.last_maintenance.strftime('%Y-%m-%d') if device.last_maintenance else "Не вказано"
                    record_device_history(device.id, current_user.id, 'update', 'Дата обслуговування', old_value, last_maintenance)
                    device.last_maintenance = last_maintenance_date
            except ValueError:
                pass  # Ігноруємо неправильний формат дати
        else:
            if device.last_maintenance:
                old_value = device.last_maintenance.strftime('%Y-%m-%d')
                record_device_history(device.id, current_user.id, 'update', 'Дата обслуговування', old_value, "Не вказано")
                device.last_maintenance = None
        
        maintenance_interval = request.form.get('maintenance_interval', type=int)
        if maintenance_interval and device.maintenance_interval != maintenance_interval:
            old_value = str(device.maintenance_interval) if device.maintenance_interval else "365"
            record_device_history(device.id, current_user.id, 'update', 'Інтервал обслуговування', old_value, str(maintenance_interval))
            device.maintenance_interval = maintenance_interval
        
        # Оновлюємо дату наступного обслуговування
        device.update_next_maintenance()
        
        # Обробляємо фінансові дані
        purchase_price = request.form.get('purchase_price', type=float)
        if purchase_price != device.purchase_price:
            old_value = str(device.purchase_price) if device.purchase_price else "Не вказано"
            record_device_history(device.id, current_user.id, 'update', 'Вартість покупки', old_value, str(purchase_price))
            device.purchase_price = purchase_price
        
        purchase_date_str = request.form.get('purchase_date', '')
        if purchase_date_str:
            try:
                purchase_date = datetime.strptime(purchase_date_str, '%Y-%m-%d').date()
                if purchase_date != device.purchase_date:
                    old_value = device.purchase_date.strftime('%Y-%m-%d') if device.purchase_date else "Не вказано"
                    record_device_history(device.id, current_user.id, 'update', 'Дата покупки', old_value, purchase_date_str)
                    device.purchase_date = purchase_date
            except ValueError:
                pass
        else:
            if device.purchase_date:
                old_value = device.purchase_date.strftime('%Y-%m-%d')
                record_device_history(device.id, current_user.id, 'update', 'Дата покупки', old_value, "Не вказано")
                device.purchase_date = None
        
        
        db.session.commit()
        flash('Пристрій успішно оновлено!')
        return redirect(url_for('devices.device_detail', device_id=device.id))
    
    # Отримуємо список міст для адміністраторів
    if current_user.is_admin:
        cities = City.query.all()
    else:
        cities = [current_user.city]
        
    return render_template('edit_device.html', device=device, cities=cities)

@devices_bp.route('/device/<int:device_id>/delete', methods=['POST'])
@login_required
def delete_device(device_id):
    device = Device.query.get_or_404(device_id)
    
    # Перевіряємо, чи має користувач доступ до цього пристрою
    if not current_user.is_admin and device.city_id != current_user.city_id:
        abort(403)
    
    # Видаляємо фотографії пристрою з файлової системи
    for photo in device.photos:
        try:
            os.remove(os.path.join(current_app.config['UPLOAD_FOLDER'], photo.filename))
        except:
            # Якщо файл не знайдено, просто продовжуємо
            pass
    
    # Записуємо історію видалення ПЕРЕД видаленням пристрою
    record_device_history(device.id, current_user.id, 'delete', device=device)
    
    # Видаляємо пристрій (каскадне видалення також видалить записи фото)
    db.session.delete(device)
    db.session.commit()
    
    flash('Пристрій успішно видалено!')
    return redirect(url_for('devices.devices'))



@devices_bp.route('/device/<int:device_id>/add_photo', methods=['POST'])
@login_required
def add_device_photo(device_id):
    # Rate limiting застосовується через глобальні обмеження в app.py
    device = Device.query.get_or_404(device_id)
    
    # Перевіряємо, чи має користувач доступ до цього пристрою
    if not current_user.is_admin and device.city_id != current_user.city_id:
        abort(403)
    
    if 'photo' in request.files:
        photo = request.files['photo']
        if photo and allowed_file(photo.filename):
            # Генеруємо унікальне ім'я файлу
            filename = secure_filename(photo.filename)
            unique_filename = f"{uuid.uuid4()}_{filename}"
            
            # Зберігаємо файл
            file_path = os.path.join(current_app.config['UPLOAD_FOLDER'], unique_filename)
            photo.save(file_path)
            
            # Оптимізуємо зображення
            optimized_path = optimize_image(file_path)
            if optimized_path and optimized_path != file_path:
                # Якщо файл було конвертовано
                unique_filename = os.path.basename(optimized_path)
                file_path = optimized_path
            
            # Генеруємо thumbnail'и
            thumbnails = generate_thumbnails(file_path)
            
            # Створюємо WebP версію (якщо браузер підтримує)
            webp_path = convert_to_webp(file_path)
            
            # Створюємо запис у базі даних
            device_photo = DevicePhoto(
                filename=unique_filename,
                original_filename=filename,
                device_id=device.id
            )
            db.session.add(device_photo)
            db.session.commit()
            
            flash('Фото успішно додано!')
    
    return redirect(url_for('devices.device_detail', device_id=device_id))

@devices_bp.route('/device/photo/<int:photo_id>/delete', methods=['POST'])
@login_required
def delete_device_photo(photo_id):
    photo = DevicePhoto.query.get_or_404(photo_id)
    device = photo.device
    
    # Перевіряємо, чи має користувач доступ до цього пристрою
    if not current_user.is_admin and device.city_id != current_user.city_id:
        abort(403)
    
    # Видаляємо файл
    try:
        os.remove(os.path.join(current_app.config['UPLOAD_FOLDER'], photo.filename))
    except:
        # Якщо файл не знайдено, просто продовжуємо
        pass
    
    # Видаляємо запис з бази даних
    db.session.delete(photo)
    db.session.commit()
    
    flash('Фото успішно видалено!')
    return redirect(url_for('devices.device_detail', device_id=device.id))

@devices_bp.route('/device/<int:device_id>/add-repair-expense', methods=['POST'])
@login_required
def add_repair_expense(device_id):
    """Додавання витрат на ремонт"""
    device = Device.query.get_or_404(device_id)
    
    # Перевіряємо доступ
    if not current_user.is_admin and device.city_id != current_user.city_id:
        abort(403)
    
    try:
        expense = RepairExpense(
            device_id=device_id,
            amount=request.form.get('amount', type=float),
            description=request.form.get('description', ''),
            repair_date=datetime.strptime(request.form['repair_date'], '%Y-%m-%d').date(),
            invoice_number=request.form.get('invoice_number', '')
        )
        db.session.add(expense)
        db.session.commit()
        
        flash('Витрати на ремонт успішно додано!', 'success')
        log_user_activity(current_user.id, f'Додано витрати на ремонт: {expense.amount} для {device.name}', request.remote_addr, request.url)
    except Exception as e:
        db.session.rollback()
        flash(f'Помилка при додаванні витрат: {str(e)}', 'error')
    
    return redirect(url_for('devices.device_detail', device_id=device_id))

@devices_bp.route('/device/repair-expense/<int:expense_id>/delete', methods=['POST'])
@login_required
def delete_repair_expense(expense_id):
    """Видалення витрат на ремонт"""
    expense = RepairExpense.query.get_or_404(expense_id)
    device = expense.device
    
    # Перевіряємо доступ
    if not current_user.is_admin and device.city_id != current_user.city_id:
        abort(403)
    
    amount = expense.amount
    db.session.delete(expense)
    db.session.commit()
    
    flash('Витрати на ремонт успішно видалено!', 'success')
    log_user_activity(current_user.id, f'Видалено витрати на ремонт: {amount} для {device.name}', request.remote_addr, request.url)
    
    return redirect(url_for('devices.device_detail', device_id=device.id))

@devices_bp.route('/uploads/<filename>')
@login_required
def uploaded_file(filename):
    # Отримуємо фото, щоб перевірити права доступу
    photo = DevicePhoto.query.filter_by(filename=filename).first_or_404()
    device = photo.device
    
    # Перевіряємо, чи має користувач доступ до цього пристрою
    if not current_user.is_admin and device.city_id != current_user.city_id:
        abort(403)
    
    # Перевіряємо чи потрібен thumbnail
    size = request.args.get('size', 'original')
    if size in ['thumb', 'medium', 'large']:
        # Шукаємо thumbnail
        base_name = os.path.splitext(filename)[0]
        ext = os.path.splitext(filename)[1]
        thumb_filename = f"{base_name}_{size}{ext}"
        thumb_path = os.path.join(current_app.config['UPLOAD_FOLDER'], thumb_filename)
        
        if os.path.exists(thumb_path):
            return send_from_directory(current_app.config['UPLOAD_FOLDER'], thumb_filename)
    
    # Перевіряємо чи браузер підтримує WebP
    accept_header = request.headers.get('Accept', '')
    if 'image/webp' in accept_header:
        # Шукаємо WebP версію
        base_name = os.path.splitext(filename)[0]
        webp_filename = f"{base_name}.webp"
        webp_path = os.path.join(current_app.config['UPLOAD_FOLDER'], webp_filename)
        
        if os.path.exists(webp_path):
            return send_from_directory(current_app.config['UPLOAD_FOLDER'], webp_filename)
        
    return send_from_directory(current_app.config['UPLOAD_FOLDER'], filename)

@devices_bp.route('/device/<int:device_id>/history')
@login_required
def device_history(device_id):
    device = Device.query.get_or_404(device_id)
    
    # Перевіряємо, чи має користувач доступ до цього пристрою
    if not current_user.is_admin and device.city_id != current_user.city_id:
        abort(403)
    
    history = DeviceHistory.query.filter_by(device_id=device_id).order_by(DeviceHistory.timestamp.desc()).all()
    
    # Налагоджувальна інформація
    current_app.logger.info(f"Device ID: {device_id}, History count: {len(history)}")
    for h in history:
        current_app.logger.info(f"History: {h.action}, {h.field}, {h.timestamp}")
    
    return render_template('device_history.html', device=device, history=history)

@devices_bp.route('/history/<int:device_id>')
@login_required
def device_history_by_id(device_id):
    """Перегляд історії пристрою за ID (навіть якщо пристрій видалений)"""
    
    # Отримуємо історію для цього device_id
    history = DeviceHistory.query.filter_by(device_id=device_id).order_by(DeviceHistory.timestamp.desc()).all()
    
    if not history:
        abort(404)
    
    # Перевіряємо права доступу на основі першого запису історії
    first_history = history[0]
    device = Device.query.get(device_id)  # Може бути None для видалених пристроїв
    
    # Якщо пристрій існує, перевіряємо права як зазвичай
    if device and not current_user.is_admin and device.city_id != current_user.city_id:
        abort(403)
    
    # Якщо пристрій видалений, створюємо об'єкт з збереженої інформації
    if not device and first_history:
        device = type('Device', (), {
            'id': device_id,
            'name': first_history.device_name,
            'inventory_number': first_history.device_inventory_number,
            'type': first_history.device_type,
            'serial_number': first_history.device_serial_number
        })()
    
    return render_template('device_history.html', device=device, history=history, is_deleted=not Device.query.get(device_id))

@devices_bp.route('/device/<int:device_id>/qrcode')
@login_required
def device_qrcode(device_id):
    device = Device.query.get_or_404(device_id)
    
    # Перевіряємо, чи має користувач доступ до цього пристрою
    if not current_user.is_admin and device.city_id != current_user.city_id:
        abort(403)
    
    # Створюємо QR-код з інформацією про пристрій
    qr_data = f"""
    ID: {device.id}
    Назва: {device.name}
    Інв. номер: {device.inventory_number}
    Тип: {device.type}
    S/N: {device.serial_number}
    Розташування: {device.location}
    Статус: {device.status}
    Місто: {device.city.name}
    """
    
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(qr_data)
    qr.make(fit=True)

    img = qr.make_image(fill_color="black", back_color="white")
    
    # Зберігаємо QR-код у буфер пам'яті
    img_buffer = io.BytesIO()
    img.save(img_buffer)
    img_buffer.seek(0)
    
    return send_file(
        img_buffer,
        mimetype='image/png',
        as_attachment=True,
        download_name=f'qrcode_{device.inventory_number}.png'
    )

@devices_bp.route('/device/<int:device_id>/print_qrcode')
@login_required
def print_device_qrcode(device_id):
    """Відображення сторінки для друку QR-коду пристрою"""
    device = Device.query.get_or_404(device_id)
    
    # Перевіряємо, чи має користувач доступ до цього пристрою
    if not current_user.is_admin and device.city_id != current_user.city_id:
        abort(403)
    
    return render_template('print_qrcode.html', device=device)

@devices_bp.route('/device/<int:device_id>/print_inventory')
@login_required
def print_inventory(device_id):
    """Відображення сторінки для друку інвентарного номера пристрою"""
    device = Device.query.get_or_404(device_id)
    
    # Перевіряємо, чи має користувач доступ до цього пристрою
    if not current_user.is_admin and device.city_id != current_user.city_id:
        abort(403)
    
    # Логуємо активність користувача
    log_user_activity(current_user.id, f'Друк інвентарного номера пристрою: {device.name}', request.remote_addr, request.url)
    
    return render_template('print_inventory.html', device=device)

@devices_bp.route('/devices/import_excel', methods=['GET', 'POST'])
@login_required
def import_excel():
    """Імпорт пристроїв з Excel файлу"""
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('Файл не вибрано!', 'error')
            return redirect(request.url)
        
        file = request.files['file']
        if file.filename == '':
            flash('Файл не вибрано!', 'error')
            return redirect(request.url)
        
        if file and file.filename.endswith(('.xlsx', '.xls')):
            try:
                # Читаємо Excel файл
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active
                
                imported_count = 0
                errors = []
                
                # Визначаємо місто для пристроїв
                if current_user.is_admin and request.form.get('city_id'):
                    city_id = request.form.get('city_id', type=int)
                    city = City.query.get(city_id)
                    city_prefix = city.name if city else ""
                else:
                    city_id = current_user.city_id
                    city_prefix = current_user.city.name if current_user.city else ""
                
                # Пропускаємо заголовок (перший рядок)
                for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    if not any(row):  # Пропускаємо порожні рядки
                        continue
                    
                    try:
                        name = str(row[0]) if row[0] else f"Пристрій {row_num}"
                        device_type = str(row[1]) if row[1] else "Не вказано"
                        serial_number = str(row[2]) if row[2] else ""
                        location = str(row[3]) if row[3] else ""
                        status = str(row[4]) if row[4] else "Активний"
                        notes = str(row[5]) if row[5] else ""
                        
                        # Перевіряємо, чи існує пристрій з таким серійним номером
                        if serial_number and Device.query.filter_by(serial_number=serial_number).first():
                            errors.append(f"Рядок {row_num}: Пристрій з серійним номером {serial_number} вже існує")
                            continue
                        
                        # Генеруємо унікальний інвентарний номер
                        inventory_number = generate_inventory_number()
                        
                        device = Device(
                            name=name,
                            type=device_type,
                            serial_number=serial_number,
                            inventory_number=inventory_number,
                            location=location,
                            status=status,
                            notes=notes,
                            city_id=city_id
                        )
                        
                        db.session.add(device)
                        db.session.flush()  # Отримуємо ID без коміту
                        
                        # Записуємо історію створення
                        record_device_history(device.id, current_user.id, 'create')
                        
                        imported_count += 1
                        
                    except Exception as e:
                        errors.append(f"Рядок {row_num}: Помилка обробки - {str(e)}")
                        continue
                
                db.session.commit()
                
                if imported_count > 0:
                    flash(f'Успішно імпортовано {imported_count} пристрої(в)!', 'success')
                    log_user_activity(current_user.id, f'Імпортовано {imported_count} пристроїв з Excel', request.remote_addr, request.url)
                
                if errors:
                    for error in errors[:10]:  # Показуємо перші 10 помилок
                        flash(error, 'warning')
                    if len(errors) > 10:
                        flash(f'... та ще {len(errors) - 10} помилок', 'warning')
                
                return redirect(url_for('devices.devices'))
                
            except Exception as e:
                flash(f'Помилка при обробці файлу: {str(e)}', 'error')
                return redirect(request.url)
        else:
            flash('Дозволені тільки Excel файли (.xlsx, .xls)!', 'error')
            return redirect(request.url)
    
    # Отримуємо список міст для адміністраторів
    if current_user.is_admin:
        cities = City.query.all()
    else:
        cities = [current_user.city]
    
    return render_template('import_excel.html', cities=cities)

@devices_bp.route('/devices/export_excel')
@login_required
def export_excel():
    """Експорт пристроїв в Excel файл"""
    # Отримуємо пристрої відповідно до прав користувача
    if current_user.is_admin:
        devices = Device.query.all()
    else:
        devices = Device.query.filter_by(city_id=current_user.city_id).all()
    
    # Створюємо новий Excel файл
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Пристрої"
    
    # Заголовки стовпців
    headers = [
        'ID', 'Назва', 'Тип', 'Серійний номер', 'Інвентарний номер',
        'Розташування', 'Статус', 'Місто', 'Дата створення', 'Примітки'
    ]
    
    # Стилі для заголовків
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Записуємо заголовки
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Записуємо дані пристроїв
    for row, device in enumerate(devices, 2):
        data = [
            device.id,
            device.name,
            device.type,
            device.serial_number,
            device.inventory_number,
            device.location,
            device.status,
            device.city.name if device.city else '',
            device.created_at.strftime('%Y-%m-%d %H:%M') if device.created_at else '',
            device.notes
        ]
        
        for col, value in enumerate(data, 1):
            cell = sheet.cell(row=row, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
    
    # Автоматично підганяємо ширину стовпців
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Зберігаємо файл у буфер пам'яті
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    
    # Генеруємо ім'я файлу з поточною датою
    filename = f'devices_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    log_user_activity(current_user.id, f'Експортовано {len(devices)} пристроїв в Excel', request.remote_addr, request.url)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
     )

@devices_bp.route('/device/<int:device_id>/export_pdf')
@login_required
def export_device_pdf(device_id):
    """Експорт інвентарної картки пристрою в PDF"""
    device = Device.query.get_or_404(device_id)
    
    # Перевіряємо, чи має користувач доступ до цього пристрою
    if not current_user.is_admin and device.city_id != current_user.city_id:
        abort(403)
    
    from utils_pdf import generate_device_pdf
    
    pdf_buffer = generate_device_pdf(device)
    
    log_user_activity(current_user.id, f'Експорт PDF пристрою: {device.name}', request.remote_addr, request.url)
    
    return send_file(
        pdf_buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'device_{device.inventory_number}.pdf'
    )

@devices_bp.route('/devices/export_pdf', methods=['POST'])
@login_required
def export_devices_bulk_pdf():
    """Масовий експорт пристроїв в PDF"""
    device_ids = request.form.getlist('device_ids', type=int)
    
    if not device_ids:
        flash('Не вибрано жодного пристрою для експорту!', 'warning')
        return redirect(url_for('devices.devices'))
    
    # Отримуємо пристрої відповідно до прав користувача
    if current_user.is_admin:
        devices = Device.query.filter(Device.id.in_(device_ids)).all()
    else:
        devices = Device.query.filter(
            Device.id.in_(device_ids),
            Device.city_id == current_user.city_id
        ).all()
    
    if not devices:
        flash('Не знайдено пристроїв для експорту!', 'error')
        return redirect(url_for('devices.devices'))
    
    from utils_pdf import generate_bulk_devices_pdf
    
    pdf_buffer = generate_bulk_devices_pdf(devices)
    
    log_user_activity(current_user.id, f'Масовий експорт PDF: {len(devices)} пристроїв', request.remote_addr, request.url)
    
    filename = f'devices_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    
    return send_file(
        pdf_buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )

@devices_bp.route('/qr-scanner')
@login_required
def qr_scanner():
    """Сторінка мобільного QR-сканера"""
    return render_template('qr_scanner.html')

@devices_bp.route('/devices/bulk-update-status', methods=['POST'])
@login_required
def bulk_update_status():
    """Масове оновлення статусу пристроїв"""
    device_ids = request.form.getlist('device_ids', type=int)
    new_status = request.form.get('new_status')
    
    if not device_ids or not new_status:
        flash('Не вибрано пристрої або не вказано статус', 'error')
        return redirect(url_for('devices.devices'))
    
    # Отримуємо пристрої відповідно до прав користувача
    if current_user.is_admin:
        devices = Device.query.filter(Device.id.in_(device_ids)).all()
    else:
        devices = Device.query.filter(
            Device.id.in_(device_ids),
            Device.city_id == current_user.city_id
        ).all()
    
    if not devices:
        flash('Пристрої не знайдено або у вас немає доступу', 'error')
        return redirect(url_for('devices.devices'))
    
    # Оновлюємо статус та записуємо історію
    updated_count = 0
    for device in devices:
        if device.status != new_status:
            old_status = device.status
            device.status = new_status
            record_device_history(device.id, current_user.id, 'update', 'Статус', old_status, new_status)
            updated_count += 1
    
    db.session.commit()
    
    flash(f'Статус оновлено для {updated_count} пристрої(в)', 'success')
    log_user_activity(current_user.id, f'Масове оновлення статусу: {updated_count} пристроїв', request.remote_addr, request.url)
    
    return redirect(url_for('devices.devices'))

@devices_bp.route('/devices/bulk-export-excel', methods=['POST'])
@login_required
def bulk_export_excel():
    """Масовий експорт вибраних пристроїв в Excel"""
    device_ids = request.form.getlist('device_ids', type=int)
    
    if not device_ids:
        flash('Не вибрано жодного пристрою для експорту!', 'warning')
        return redirect(url_for('devices.devices'))
    
    # Отримуємо пристрої
    if current_user.is_admin:
        devices = Device.query.filter(Device.id.in_(device_ids)).all()
    else:
        devices = Device.query.filter(
            Device.id.in_(device_ids),
            Device.city_id == current_user.city_id
        ).all()
    
    if not devices:
        flash('Пристрої не знайдено', 'error')
        return redirect(url_for('devices.devices'))
    
    # Створюємо Excel файл (використовуємо існуючу функцію)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Обрані пристрої"
    
    # Заголовки
    headers = ['ID', 'Назва', 'Тип', 'Серійний номер', 'Інвентарний номер',
               'Розташування', 'Статус', 'Місто', 'Дата створення', 'Примітки']
    
    # Стилі
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Записуємо заголовки
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Записуємо дані
    for row, device in enumerate(devices, 2):
        data = [
            device.id,
            device.name,
            device.type,
            device.serial_number,
            device.inventory_number,
            device.location,
            device.status,
            device.city.name if device.city else '',
            device.created_at.strftime('%Y-%m-%d %H:%M') if device.created_at else '',
            device.notes
        ]
        
        for col, value in enumerate(data, 1):
            cell = sheet.cell(row=row, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
    
    # Автоматично підганяємо ширину
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Зберігаємо в буфер
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    
    filename = f'selected_devices_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    log_user_activity(current_user.id, f'Масовий експорт Excel: {len(devices)} пристроїв', request.remote_addr, request.url)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@devices_bp.route('/devices/bulk_print_inventory', methods=['GET', 'POST'])
@login_required
def bulk_print_inventory():
    """Масовий друк інвентарних номерів пристроїв"""
    if request.method == 'POST':
        device_ids = request.form.getlist('device_ids', type=int)
        
        if not device_ids:
            flash('Не вибрано жодного пристрою для друку!', 'warning')
            return redirect(url_for('devices.devices'))
        
        # Отримуємо пристрої відповідно до прав користувача
        if current_user.is_admin:
            devices = Device.query.filter(Device.id.in_(device_ids)).all()
        else:
            devices = Device.query.filter(
                Device.id.in_(device_ids),
                Device.city_id == current_user.city_id
            ).all()
        
        if not devices:
            flash('Не знайдено пристроїв для друку!', 'error')
            return redirect(url_for('devices.devices'))
        
        log_user_activity(current_user.id, f'Масовий друк {len(devices)} пристроїв', request.remote_addr, request.url)
        
        return render_template('bulk_print_inventory.html', devices=devices)
    
    # GET запит - показуємо сторінку вибору пристроїв
    # Отримуємо пристрої відповідно до прав користувача
    if current_user.is_admin:
        devices = Device.query.all()
        cities = City.query.all()
    else:
        devices = Device.query.filter_by(city_id=current_user.city_id).all()
        cities = [current_user.city]
    
    return render_template('bulk_print_select.html', devices=devices, cities=cities)


@devices_bp.route('/maintenance/confirm/<int:device_id>', methods=['POST'])
@login_required
def confirm_maintenance(device_id):
    abort(404)